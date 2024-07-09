from flask import Flask, request, jsonify
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials, storage
import os

def download_excel_from_firebase(file_path):
    bucket = storage.bucket()
    blob = bucket.blob('data.xlsx')
    blob.download_to_filename(file_path)

def upload_excel_to_firebase(file_path):
    bucket = storage.bucket()
    blob = bucket.blob('data.xlsx')
    blob.upload_from_filename(file_path)

def copy_cell_styles(source_cell, target_cell):
    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = source_cell.font.copy()
        if source_cell.border:
            target_cell.border = source_cell.border.copy()
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.protection:
            target_cell.protection = source_cell.protection.copy()
        if source_cell.alignment:
            target_cell.alignment = source_cell.alignment.copy()

def is_valid_row(row):
    # 한 열이라도 내용이 있는지 확인
    return any(cell.value for cell in row)

def process_and_update_excel(existing_file_path, new_file_path):
    try:
        # Load existing and new workbooks
        existing_wb = load_workbook(existing_file_path)
        new_wb = load_workbook(new_file_path)
    except Exception as e:
        return False, f"Error loading Excel files: {e}"
    
    existing_ws = existing_wb.active
    new_ws = new_wb.active

    # 검증된 헤더 목록 (식별번호 제거)
    expected_headers = ['총판', '대행사', '셀러', '메인 키워드', '서브 키워드', '상품 URL', 'MID값', '원부 URL', '원부 MID값', '시작일', '종료일', '유입수']
    
    # 헤더 검증
    headers = [cell.value if cell.value is not None else '' for cell in new_ws[1]]
    if headers != expected_headers:
        return False, "올바른 형식의 파일이 아닙니다."

    # Get current max identifier
    current_max_identifier = 0
    # 기존 파일에서 유효한 행만 남김
    valid_existing_rows = []
    for row in existing_ws.iter_rows(min_row=2, values_only=False):
        if is_valid_row(row):
            valid_existing_rows.append(row)
            current_max_identifier += 1

    # 기존 데이터 초기화
    existing_ws.delete_rows(2, existing_ws.max_row)

    # 유효한 기존 행 추가
    for row_index, row in enumerate(valid_existing_rows, start=2):
        for col_index, cell in enumerate(row, start=1):
            new_cell = existing_ws.cell(row=row_index, column=col_index)
            new_cell.value = cell.value if cell.value is not None else None
            copy_cell_styles(cell, new_cell)

    # 새로운 데이터 추가
    for row in new_ws.iter_rows(min_row=2, values_only=False):
        if not is_valid_row(row):
            continue
        current_max_identifier += 1
        for col_index, cell in enumerate(row, start=1):  # 식별번호 열 추가
            new_cell = existing_ws.cell(row=current_max_identifier, column=col_index+1)
            new_cell.value = cell.value if cell.value is not None else None
            copy_cell_styles(cell, new_cell)
        
        print(f"Row {current_max_identifier} added with values {[cell.value for cell in row]}")

        # 식별번호 추가
        existing_ws.cell(row=current_max_identifier, column=1).value = current_max_identifier

    # 디버깅을 위한 데이터 확인
    print("Saving updated Excel file...")
    for row in existing_ws.iter_rows(values_only=True):
        print(row)

    existing_wb.save(existing_file_path)
    print("Excel file saved successfully.")
    return True, "파일이 성공적으로 업로드되었습니다."